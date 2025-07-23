from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
import pandas as pd

# Step 1: Load formatting from File A
wb_a = load_workbook("file_A.xlsx")
ws_a = wb_a.active

# Extract header (row 1) and first data row (row 2) formatting
header_styles = []
data_styles = []
column_widths = {}

for col in range(1, ws_a.max_column + 1):
    col_letter = get_column_letter(col)
    column_widths[col_letter] = ws_a.column_dimensions[col_letter].width

    # Header cell (row 1)
    header_cell = ws_a.cell(row=1, column=col)
    header_styles.append(
        {
            "alignment": header_cell.alignment,
            "font": header_cell.font,
            "fill": header_cell.fill,
        }
    )

    # Data cell (row 2)
    data_cell = ws_a.cell(row=2, column=col)
    data_styles.append(
        {
            "alignment": data_cell.alignment,
            "font": data_cell.font,
            "fill": data_cell.fill,
            "number_format": data_cell.number_format,
        }
    )

# Step 2: Process your data in pandas (e.g., merge A and B)
# Here we just make an example DataFrame
df = pd.DataFrame({"Product": ["A", "B"], "Quantity": [10, 20], "Price": [2.5, 3.0]})

# Step 3: Write to new Excel and apply formatting
wb_new = Workbook()
ws_new = wb_new.active

# Apply column widths
for col in range(1, len(df.columns) + 1):
    col_letter = get_column_letter(col)
    if col_letter in column_widths:
        ws_new.column_dimensions[col_letter].width = column_widths[col_letter]

# Write headers with formatting
for col_idx, col_name in enumerate(df.columns, start=1):
    cell = ws_new.cell(row=1, column=col_idx, value=col_name)
    if col_idx <= len(header_styles):
        style = header_styles[col_idx - 1]
        cell.font = style["font"]
        cell.alignment = style["alignment"]
        cell.fill = style["fill"]

# Write data rows with formatting
for row_idx, row in enumerate(df.itertuples(index=False), start=2):
    for col_idx, value in enumerate(row, start=1):
        cell = ws_new.cell(row=row_idx, column=col_idx, value=value)
        if col_idx <= len(data_styles):
            style = data_styles[col_idx - 1]
            cell.font = style["font"]
            cell.alignment = style["alignment"]
            cell.fill = style["fill"]
            cell.number_format = style["number_format"]

# Save final file
wb_new.save("merged_styled.xlsx")

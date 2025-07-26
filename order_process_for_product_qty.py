import pandas as pd
import re
from natsort import natsort_keygen, ns
from copy import copy
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


def get_order_completed_df(order_completed_filepath_list, income_released_order_ids):

    df_to_merge = []
    for order_completed in order_completed_filepath_list:
        order_completed_df = pd.read_excel(
            order_completed, sheet_name="orders", header=0
        )
        """Only remove refunded orders in Order.completed excel file. Example (from May's income released):
            Order ID 2505065PU181Y4 has 2 items in Order.completed excel file:
                COLD_PACK_STRAP	    Black Strap Only
                COLD_PACK_6INCHES	Blue 6 Inches
            
            Say if COLD_PACK_STRAP is refunded, in the Income.released excel file, 2505065PU181Y4 will have a refund id. If this order id is removed because 
            'refund id' column in Income.released is not empty, then wont be able to find the remaining orders in Order.completed.

            Remove refunded orders in Order.completed, as each unique item with the same Order ID has different rows. Can safely remove the data rows which are
            refunded
        """

        # Only want to get product quantity for non-refunded orders, to decrease the stock count
        remove_refunded_orders_df = order_completed_df.loc[
            order_completed_df["Return / Refund Status"].isna()
        ]

        # Only the the orders that are paid by shopee, in the Order.completed df
        released_order_completed_df = remove_refunded_orders_df[
            remove_refunded_orders_df["Order ID"].isin(income_released_order_ids)
        ]

        final_order_completed_df = released_order_completed_df[
            [
                "Product Name",
                "SKU Reference No.",
                "Variation Name",
                "Quantity",
            ]
        ].copy()

        # For debug purpose, can check whether different Order.completed files are properly merged
        final_order_completed_df["From"] = "_".join(
            re.findall(r"\d{8}", order_completed)
        )
        # Avoid nan values affect the groupby.sum()
        final_order_completed_df["Variation Name"] = final_order_completed_df[
            "Variation Name"
        ].fillna("No Variant")

        final_order_completed_df["Quantity"] = pd.to_numeric(
            final_order_completed_df["Quantity"]
        )

        df_to_merge.append(final_order_completed_df)

    merged_order_completed = pd.concat(df_to_merge, ignore_index=True)

    return merged_order_completed


def get_product_quantity(merged_order_completed):
    """Sorting
    natural sort, where digits and text are sorted in a way that make sense to human. Example after sorted:

        COLD_PACK_90GRM	    C (90gram) [Insulin]
        COLD_PACK_280GRM	B (280gram)
        COLD_PACK_600GRM	A (600gram)

    Sort by numbers 90, 280, 600. But why C B A doesnt get sorted? Because .sort_values() act in priority order.
    .sort_values(
            [
                "Product Name",
                "SKU Reference No.",
                "Variation Name",
            ]

    Means that sort by product name, when no order to sort -> sort by sku, if there is a tie -> sort by variation name.

    ns.IGNORECASE make the sorting case-insensitive so that, if not order would be affected. Example:

        FORA ....
        Finger ....

    this happens because in ASCII, "O" (79) which is smaller than "i" (105)

    Why remove _ ?
    For case like:
        MULTI100
        MULTI_30

    The sort just don't work properly without removing underscore and other special characters.
    """
    natsort_func = natsort_keygen(alg=ns.IGNORECASE)

    # Not removing "." as there might be decimal number. Said Black sport bandage 2.5, remove it will affect sorting result
    remove_non_alphabets = lambda s: re.sub(r"[^a-zA-Z0-9. ]", "", s)

    product_qty_df = (
        merged_order_completed.groupby(
            [
                "Product Name",
                "SKU Reference No.",
                "Variation Name",
            ]
        )["Quantity"]
        .sum()
        .reset_index()
        .sort_values(
            [
                "Product Name",
                "SKU Reference No.",
                "Variation Name",
            ],
            key=lambda x: x.map(remove_non_alphabets).map(natsort_func),
        )
        .reset_index(drop=True)
    )
    # Under same product name, remove duplicated SKU
    product_qty_df["SKU Reference No."] = product_qty_df.groupby(
        ["Product Name", "SKU Reference No."]
    )["SKU Reference No."].transform(lambda x: x.mask(x.duplicated()))

    product_qty_df["Product Name"] = product_qty_df["Product Name"].mask(
        product_qty_df["Product Name"].duplicated()
    )
    return product_qty_df


# Save the cell's format and width of the Order.completed excel file, to prevent the generated Product.quantity excel file has no format.
def get_order_completed_format(any_order_completed_file_path):
    wb = load_workbook(any_order_completed_file_path)
    ws = wb.active

    # letter key is the column alphabet in the Order.completed.excel file
    col_format = {
        "Product Name": {
            "letter": "M",
            "width": None,
            "header_style": None,
            "data_style": None,
        },
        "SKU Reference No.": {
            "letter": "N",
            "width": None,
            "header_style": None,
            "data_style": None,
        },
        "Variation Name": {
            "letter": "O",
            "width": None,
            "header_style": None,
            "data_style": None,
        },
        "Quantity": {
            "letter": "R",
            "width": None,
            "header_style": None,
            "data_style": None,
        },
    }
    # Use copy() for the format values if not this workbook is active only in this function, the cells are accessed by reference.
    # After this function is executed, the reference is gone.
    for format in col_format.values():
        format["width"] = ws.column_dimensions[format["letter"]].width
        header_row = ws.cell(row=1, column=column_index_from_string(format["letter"]))
        format["header_style"] = {
            "font": copy(header_row.font),
            "fill": copy(header_row.fill),
            "alignment": copy(header_row.alignment),
        }

        data_row = ws.cell(row=2, column=column_index_from_string(format["letter"]))
        format["data_style"] = {
            "font": copy(data_row.font),
            "fill": copy(data_row.fill),
            "alignment": copy(data_row.alignment),
        }

    return col_format


# Generated a Product.quantity excel file based on the Income.released excel file
def save_product_quantity(
    order_completed_format_dict, product_qty_df, income_released_filename
):
    wb = Workbook()
    ws = wb.active

    for idx, col_name in enumerate(product_qty_df.columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = (
            order_completed_format_dict[col_name]["width"]
        )
        header_row = ws.cell(row=1, column=idx, value=col_name)
        header_row.font = order_completed_format_dict[col_name]["header_style"]["font"]
        header_row.fill = order_completed_format_dict[col_name]["header_style"]["fill"]
        header_row.alignment = order_completed_format_dict[col_name]["header_style"][
            "alignment"
        ]

    # Draw a line for each unique product name, for better visual clarity if needed to be print
    cell_bottom_underline = Border(bottom=Side(style="thin"))

    for r_id, row_data_tuple in enumerate(
        product_qty_df.itertuples(index=False), start=2
    ):
        for c_id, col_value in enumerate(row_data_tuple, start=1):
            data_row = ws.cell(row=r_id, column=c_id, value=col_value)
            data_row.font = order_completed_format_dict[col_name]["data_style"]["font"]
            data_row.fill = order_completed_format_dict[col_name]["data_style"]["fill"]
            data_row.alignment = order_completed_format_dict[col_name]["data_style"][
                "alignment"
            ]

            # Don't draw line when the next row of "Product Name" column is nan
            """Example:  
                        "Product Name"
            row237      Fiber Drink.....
            row238
                            "Skip"     (No line Drawn because row 239 is empty)
            row239
                        "-------------"(row240 has value)
            row240      FORA...
            """

            if (
                r_id <= len(product_qty_df)
                and c_id == 1
                and pd.isna(product_qty_df.iloc[r_id - 1]["Product Name"])
            ):
                continue

            data_row.border = cell_bottom_underline

    # Style for the last row, which shows the sums of all the product quantity
    light_blue_fill = PatternFill(fill_type="solid", fgColor="ADD8E6")
    bold_font = Font(bold=True)
    left_alignment = Alignment(horizontal="left")

    last_row = r_id + 1

    # Only the first and the last column has value
    for cells in ws[f"A{last_row}:D{last_row}"]:
        for cell in cells:
            if cell.coordinate == f"A{last_row}":
                cell.value = "Grand total"
            elif cell.coordinate == f"D{last_row}":
                cell.value = product_qty_df["Quantity"].sum()
                cell.alignment = left_alignment

            cell.fill = light_blue_fill
            cell.font = bold_font

    product_qty_filename = f"Product.quantity.{income_released_filename}.xlsx"
    wb.save(product_qty_filename)
    return product_qty_filename
